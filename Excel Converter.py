import pandas as pd
from openpyxl import load_workbook
import os
import tkinter as tk
from tkinter import *
from tkinter import filedialog
from openpyxl.utils.dataframe import dataframe_to_rows
import json
import sys

# Created by Michael Carlson, 8-2-24

# Global vars
# Dictionary to store mapping of columns and path to store settings
mapping = {}
user_path = os.getenv("LOCALAPPDATA") + "\\Excel Transfer Settings"
is_error = False
# Files collected
source_file = ""
template_file = ""
output_path = ""
errors_occurred = 0


# Function to read Excel file
def read_excel(file_path):
    global errors_occurred
    # Try to read Excel file and if error, show popup
    try:
        return pd.read_excel(file_path)
    except:
        # Only show error message once
        if errors_occurred < 1:
            # Error pop up
            err = tk.Tk()
            err.title("File Error")
            err.minsize(300, 100)
            err.lift()
            err.attributes("-topmost", True)
            err.attributes("-topmost", False)
            err_directions = Label(
                text="Problem selecting file. Make sure the selected file is an Excel file(.xlsx). "
                     "Closing application.",
                wraplength=260, padx=20, pady=10)
            err_directions.grid(row=0, column=0)

            # Close button
            err_close_button = Button(text="Okay", command=err.destroy)
            err_close_button.grid(row=1, column=0, pady=10)
            err.mainloop()

        # Since error occurred, set run error function
        error_occur()
        # Add 1 to error counter
        errors_occurred += 1


# If error occurs, set is_error to True to be caught later
def error_occur():
    global is_error
    is_error = True


# Function to merge data based on mappings
def merge_data(template_data_frame, source_data_frame, column_mappings):
    # Loops through the column mapping and assigns source columns to template columns
    for target_col, source_col in column_mappings.items():
        template_data_frame[target_col] = source_data_frame[source_col]
    return template_data_frame


# Function to write DataFrame to Excel file with the same formatting
def write_excel_with_template(template_path, df, out_path):
    # Ensure the directory exists
    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    # Load the template workbook
    book = load_workbook(template_path)
    sheet = book.active

    # Goes through every row and every cell in the row, copying the format over
    for row_index, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for col_index, value in enumerate(row, 1):
            # Gets a specific cell and its formating
            cell = sheet.cell(row=row_index, column=col_index, value=value)

    # Save the workbook
    if not out_path.endswith(".xlsx"):
        out_path += ".xlsx"
    book.save(out_path)


# Write Excel file and success pop up
def end_script():
    # Merge dataframes and create Excel file
    merged_df = merge_data(template_df, source_df, mapping)
    write_excel_with_template(template_file, merged_df, output_path)

    # Success window
    success = tk.Tk()
    success.title("Success!")
    success.minsize(200, 100)
    success.lift()
    success.attributes("-topmost", True)
    success.attributes("-topmost", False)
    success.config(bg="#CE1126")
    message = Label(success, text="File created Successfully!", bg="#CE1126", font=("Arial", 14), fg="white",
                    wraplength=260, padx=20, pady=10)
    message.grid(row=0, column=0)

    # Success button
    message_button = Button(success, text="Okay", command=success.destroy)
    message_button.grid(row=1, column=0, pady=10)
    success.mainloop()
    return


# Add selected listbox to entry
def add_entry():
    # Clear entry boxes(Entry boxes were part of old system, but now just used to hold data)
    template_name_entry.delete(0, tk.END)
    source_index_entry.delete(0, tk.END)
    # Add to entry without index, if no selection, do nothing
    temp_sel = template_listbox.curselection()
    sour_sel = source_listbox.curselection()
    if not temp_sel and not sour_sel:
        return
    template_name_entry.insert(0, template_listbox.get(temp_sel)[3:])
    source_index_entry.insert(0, source_listbox.get(sour_sel)[3:])
    # Add mapping to dict
    user_template_string = template_name_entry.get()
    user_source_string = source_index_entry.get()
    mapping[user_template_string] = user_source_string
    # Add mapping to listbox
    mappings_listbox.insert(mappings_listbox.size(), str(mappings_listbox.size()) + ": " + user_source_string
                            + " > " + user_template_string)


# Delete selected mapping
def delete_map():
    # Delete mapping from listbox
    selection = mappings_listbox.curselection()
    if not selection:
        return
    to_be_del = mappings_listbox.get(selection)
    to_be_del_index = int(to_be_del[0])
    mappings_listbox.delete(to_be_del_index)

    # Get template column name
    template_del = (to_be_del.split(" > ", 1)[1])

    # Delete mapping from dict
    del mapping[template_del]
    # Clear listbox and put the rest of the mappings back in
    mappings_listbox.delete(0, END)
    temp_counter = 0
    for count in mapping:
        mappings_listbox.insert(temp_counter, str(temp_counter) + ": " + str({mapping[count]})[2:-2] + " > "
                                + str({count})[2:-2])
        temp_counter = temp_counter + 1


# Saves the file
def save_file():
    # Make save path open file explorer
    save_path = filedialog.asksaveasfilename(title="Output File Name and Location", filetypes=[("Json", ".json")],
                                             initialdir=user_path)

    # Creates json file and dumps info into it
    json_dict = json.dumps(mapping)
    with open(save_path + ".json", "w") as saved_json:
        json.dump(json_dict, saved_json)


# Loads the selected json file
def load_file():
    # Gets the mapping dict and path to where settings are stored
    global mapping
    global user_path
    selected_json = filedialog.askopenfilename(initialdir=user_path, title="Load Mapping",
                                               filetypes=[("Json", "*.json")])

    # If 'cancel' was selected, json will be empty
    if not selected_json:
        return

    # Opens json and reads it to data
    with open(selected_json, "r") as dict_saved:
        data = json.load(dict_saved)
        # Turns json back to dict
        dict_data = json.loads(data)
        mapping = dict_data
        # Puts the dict into the listbox
        temp_counter = 0
        for j in mapping:
            mappings_listbox.insert(temp_counter,
                                    str(temp_counter) + ": " + str({mapping[j]})[2:-2] + " mapped to "
                                    + str({j})[2:-2])
            temp_counter = temp_counter + 1


# Opens file explorer to select source file
def source_file_open():
    global source_file
    global source_file_display
    try:
        # Sets state to allow insert and then disables to stop editing
        source_file_display.config(state=NORMAL)
        source_file_display.delete(0, END)
        source_file = filedialog.askopenfilename(title="Source File", filetypes=[("Excel Files", ".xlsx .xls")])
        source_file_display.insert(0, os.path.basename(source_file))
        source_file_display.config(state=DISABLED)
        # if not source_file:
        #     sys.exit()
    except:
        sys.exit()


# Opens file explorer to select template file
def template_file_open():
    global template_file
    global temp_file_display
    try:
        # Sets state to allow insert and then disables to stop editing
        temp_file_display.config(state=NORMAL)
        temp_file_display.delete(0, END)
        template_file = filedialog.askopenfilename(title="Template File", filetypes=[("Excel Files", ".xlsx .xls")])
        temp_file_display.insert(0, os.path.basename(template_file))
        temp_file_display.config(state=DISABLED)
        # if not template_file:
        #     sys.exit()
    except:
        sys.exit()


# Opens file explorer to select template file
def mapping_file_open():
    global output_path
    global mapping_file_display
    try:
        # Sets state to allow insert and then disables to stop editing
        mapping_file_display.config(state=NORMAL)
        mapping_file_display.delete(0, END)
        output_path = filedialog.asksaveasfilename(title="Output File Name and Location",
                                                   filetypes=[("Excel Files", ".xlsx .xls")])
        mapping_file_display.insert(0, os.path.basename(output_path))
        mapping_file_display.config(state=DISABLED)
        # if not output_path:
        #     sys.exit()
    except:
        sys.exit()


# Closes the program
def on_close():
    sys.exit()


# Initial prompt telling users how to  use the program
prompt = tk.Tk()
prompt.title("Directions")
prompt.minsize(760, 380)
prompt.config(bg="#CE1126")
directions = Label(text="Use the buttons below to open file explorers and select your files.\n\n"
                        " •  First, select your source file, information is drawn from here\n"
                        " •  Next, select your template Excel file, information is placed in here\n"
                        " •  Finally, select the location of the new file being created and give it a name\n\n"
                        "You may write over an existing file if needed.\n", font=("Arial", 14),
                   wraplength=640, padx=20, pady=10, bg="#CE1126", fg="#FFFFFF", justify=LEFT)
directions.grid(row=1, column=1)

# Source file label, button, and entry
source_file_label = Label(text="Select Source File", font=("Arial", 14), bg="#CE1126", fg="#FFFFFF", justify=LEFT)
source_file_label.grid(row=3, column=1, sticky=W)
source_file_button = Button(text="Select File", command=source_file_open)
source_file_button.grid(row=3, column=3, padx=10)
source_file_display = Entry(font=("Arial", 14), bg="#CE1126", fg="#FFFFFF", state=DISABLED,
                            disabledbackground="#CE1126", disabledforeground="#FFFFFF")
source_file_display.grid(row=5, column=1, sticky=W, padx=15)

# Template file label, button, and entry
temp_file_label = Label(text="Select Template File", font=("Arial", 14), bg="#CE1126", fg="#FFFFFF", justify=LEFT)
temp_file_label.grid(row=7, column=1, sticky=W)
temp_file_button = Button(text="Select File", command=template_file_open)
temp_file_button.grid(row=7, column=3, padx=10)
temp_file_display = Entry(font=("Arial", 14), bg="#CE1126", fg="#FFFFFF", state=DISABLED, disabledbackground="#CE1126",
                          disabledforeground="#FFFFFF")
temp_file_display.grid(row=9, column=1, sticky=W, padx=15)

# Template file label, button, and entry
mapping_file_label = Label(text="Select Mapping File Name and Location", font=("Arial", 14), bg="#CE1126", fg="#FFFFFF",
                           justify=LEFT)
mapping_file_label.grid(row=11, column=1, sticky=W)
mapping_file_button = Button(text="Select File", command=mapping_file_open)
mapping_file_button.grid(row=11, column=3, padx=10)
mapping_file_display = Entry(font=("Arial", 14), bg="#CE1126", fg="#FFFFFF", state=DISABLED,
                             disabledbackground="#CE1126", disabledforeground="#FFFFFF")
mapping_file_display.grid(row=13, column=1, sticky=W, padx=15)

# Grid weights
prompt.columnconfigure(1, weight=5)
prompt.columnconfigure(3, weight=5)

prompt.rowconfigure(1, weight=10)
prompt.rowconfigure(3, weight=5)
prompt.rowconfigure(5, weight=5)
prompt.rowconfigure(7, weight=5)
prompt.rowconfigure(9, weight=5)
prompt.rowconfigure(11, weight=5)
prompt.rowconfigure(13, weight=5)

prompt.columnconfigure(0, weight=20)
prompt.columnconfigure(2, weight=5)
prompt.columnconfigure(4, weight=20)

prompt.rowconfigure(0, weight=20)
prompt.rowconfigure(2, weight=5)
prompt.rowconfigure(4, weight=2)
prompt.rowconfigure(6, weight=5)
prompt.rowconfigure(8, weight=2)
prompt.rowconfigure(10, weight=5)
prompt.rowconfigure(12, weight=2)
prompt.rowconfigure(14, weight=5)
prompt.rowconfigure(16, weight=20)


# Close button
close_button = Button(text="Proceed", command=prompt.destroy)
close_button.grid(row=15, column=1, pady=10)

# Ensures when the window is closed, the program stops and no error occurs
prompt.protocol("WM_DELETE_WINDOW", on_close)

# Main loop
prompt.mainloop()

# If no folder is there to store settings, create it
if not os.path.exists(user_path):
    os.makedirs(user_path)

# Read excel files to df
source_df = read_excel(source_file)
template_df = read_excel(template_file)

# If no output path was selected, give error
if not output_path:
    read_excel(output_path)
    error_occur()

# If error occurred while reading file, exit
if is_error:
    sys.exit()

# Create root tkinter window
root = tk.Tk()
root.title("Excel to Excel Converter")
root.geometry("1000x500")
root.minsize(600, 300)
root.lift()
root.attributes("-topmost", True)
root.attributes("-topmost", False)
root.config(bg="#CE1126")

# List box labels
template_label = Label(root, text="Template Columns", bg="#CE1126", fg="#FFFFFF")
template_label.grid(row=1, column=1, pady=10)
template_label.config(font=("Arial", 20, "bold"))
source_label = Label(root, text="Source Columns", bg="#CE1126", fg="#FFFFFF")
source_label.grid(row=1, column=3, pady=10)
source_label.config(font=("Arial", 20, "bold"))
map_label = Label(root, text="Mapping", bg="#CE1126", fg="#FFFFFF")
map_label.grid(row=1, column=5, pady=10)
map_label.config(font=("Arial", 20, "bold"))

# Configuring weights of rows with items
root.rowconfigure(1, weight=10)
root.rowconfigure(3, weight=80)
root.rowconfigure(5, weight=10)

# Configuring weights of columns with items
root.columnconfigure(1, weight=20)
root.columnconfigure(3, weight=20)
root.columnconfigure(5, weight=40)

# Configuring weights of rows with empty space
root.rowconfigure(0, weight=2)
root.rowconfigure(2, weight=2)
root.rowconfigure(4, weight=1)
root.rowconfigure(6, weight=2)
root.rowconfigure(8, weight=10)

# Configuring weights of columns with empty space
root.columnconfigure(0, weight=5)
root.columnconfigure(2, weight=3)
root.columnconfigure(4, weight=3)
root.columnconfigure(6, weight=5)

# Create 3 frames for list boxes and configure weights
temp_frame = Frame(root)
temp_frame.grid(row=3, column=1, padx=30, sticky=N+S+E+W)
temp_frame.columnconfigure(0, weight=100)
temp_frame.columnconfigure(1, weight=1)
temp_frame.rowconfigure(0, weight=80)

source_frame = Frame(root)
source_frame.grid(row=3, column=3, sticky=N+S+E+W)
source_frame.columnconfigure(0, weight=100)
source_frame.columnconfigure(1, weight=1)
source_frame.rowconfigure(0, weight=80)

mapping_frame = Frame(root)
mapping_frame.grid(row=3, column=5, padx=30, sticky=N+S+E+W)
mapping_frame.columnconfigure(0, weight=100)
mapping_frame.columnconfigure(1, weight=1)
mapping_frame.rowconfigure(0, weight=80)

# create the template listbox
template_listbox = Listbox(temp_frame, selectmode=SINGLE, exportselection=False, width=8, height=4, bg="#FFFFFF",
                           relief=GROOVE)
template_listbox.config(font=("Arial", 12))

# Put df columns in list box
for i in range(template_df.columns.size):
    template_listbox.insert(i, str(i) + ": " + template_df.columns[i])

# Scrollbar
temp_scroll = Scrollbar(temp_frame, orient=VERTICAL)
template_listbox.config(yscrollcommand=temp_scroll.set)
temp_scroll.config(command=template_listbox.yview)
temp_scroll.grid(column=1, sticky=E+N+S)

# Put listbox in grid
template_listbox.grid(row=0, column=0, sticky=N+S+E+W)

# create the source listbox
source_listbox = Listbox(source_frame, selectmode=SINGLE, exportselection=False, width=8, height=4, bg="#FFFFFF",
                         relief=GROOVE)

# Put df columns in list box
for i in range(source_df.columns.size):
    source_listbox.insert(i, str(i) + ": " + source_df.columns[i])

# Scrollbar
source_scroll = Scrollbar(source_frame, orient=VERTICAL)
source_listbox.config(yscrollcommand=source_scroll.set)
source_scroll.config(command=source_listbox.yview)
source_scroll.grid(column=1, sticky=E+N+S)

# Put listbox in grid
source_listbox.grid(row=0, column=0, sticky=N+S+E+W)
source_listbox.config(font=("Arial", 12))

# Create the mappings list box
mappings_listbox = Listbox(mapping_frame, selectmode=SINGLE, exportselection=False, width=20, height=4, bg="#FFFFFF",
                           relief=GROOVE)

# Scrollbar
map_scroll = Scrollbar(mapping_frame, orient=VERTICAL)
mappings_listbox.config(yscrollcommand=map_scroll.set)
map_scroll.config(command=mappings_listbox.yview)
map_scroll.grid(column=1, sticky=E+N+S)

# Put listbox in grid
mappings_listbox.grid(row=0, column=0, sticky=N+S+E+W)
mappings_listbox.config(font=("Arial", 12))

# Entry boxes
template_name_entry = tk.Entry(root)
source_index_entry = tk.Entry(root)

# Save button
save_button = tk.Button(root, text="Save Current Mapping", command=save_file)
save_button.config(font=("Arial", 12))
save_button.grid(row=5, column=1)

# Load button
load_button = tk.Button(root, text="Load Mapping", command=load_file)
load_button.config(font=("Arial", 12))
load_button.grid(row=7, column=1)

# Add selected to entry box button
entry_button = tk.Button(root, text="Add", command=add_entry)
entry_button.config(font=("Arial", 12))
entry_button.grid(row=5, column=3)

# End script button
end_button = tk.Button(root, text="Create Excel File", command=end_script)
end_button.config(font=("Arial", 12))
end_button.grid(row=7, column=3)

# Delete button
delete_button = tk.Button(root, text="Delete mapping", command=delete_map)
delete_button.config(font=("Arial", 12))
delete_button.grid(row=5, column=5)

# Root main loop
root.mainloop()
